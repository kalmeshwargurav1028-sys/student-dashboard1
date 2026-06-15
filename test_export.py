from app import app, export_attendance_snapshot, db
import datetime
import io
import openpyxl

# We'll simulate what happens inside export_attendance_snapshot
with app.test_request_context('/teacher/daily_report/export'):
    today_date = datetime.datetime.now().strftime('%Y-%m-%d')
    att_doc = db.attendance.find_one({'date': today_date})
    records = att_doc.get('records', {}) if att_doc else {}
    
    students = list(db.students.find({}, {'id': 1, 'student_class': 1}))
    class_stats = {}
    for s in students:
        c_name = s.get('student_class')
        if not c_name: continue
        
        if c_name not in class_stats:
            class_stats[c_name] = {'total': 0, 'present': 0}
            
        class_stats[c_name]['total'] += 1
        status = records.get(s.get('id'))
        if status in ['Present', 'Late']:
            class_stats[c_name]['present'] += 1
                
    snapshot = []
    for c_name, stats in class_stats.items():
        snapshot.append({
            'class_name': c_name,
            'total': stats['total'],
            'present': stats['present'],
            'absent': stats['total'] - stats['present']
        })
    snapshot.sort(key=lambda x: x['class_name'])
    print("Snapshot created successfully:", snapshot)
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Snapshot"
    
    headers = ["Class", "Total Students", "Present", "Absent"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
    for row_num, item in enumerate(snapshot, 2):
        ws.cell(row=row_num, column=1, value=item['class_name'])
        ws.cell(row=row_num, column=2, value=item['total']).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=3, value=item['present']).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=item['absent']).alignment = Alignment(horizontal="center")
        
        ws.cell(row=row_num, column=3).font = Font(color="059669", bold=True)
        ws.cell(row=row_num, column=4).font = Font(color="DC2626", bold=True)
        
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    print("Excel saved successfully")
